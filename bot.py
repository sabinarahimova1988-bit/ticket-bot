import os
import json
import logging
import base64
import xml.etree.ElementTree as ET
from datetime import datetime
import imaplib
import email
from email.header import decode_header
import urllib.request
from anthropic import AsyncAnthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, Bot
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import asyncio
import psycopg2
import psycopg2.extras

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

ANTHROPIC_CLIENT = AsyncAnthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
ALLOWED_USER_ID = int(os.environ.get("ALLOWED_USER_ID", "0"))
GMAIL_USER = os.environ.get("GMAIL_USER", "")
GMAIL_PASSWORD = os.environ.get("GMAIL_PASSWORD", "")

DATABASE_URL = os.environ.get("DATABASE_URL", "")

def get_db():
    return psycopg2.connect(DATABASE_URL)

def init_db():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS tickets (
                    id BIGINT PRIMARY KEY,
                    data JSONB NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS seen_emails (
                    email_id TEXT PRIMARY KEY
                )
            """)
        conn.commit()

TICKET_KEYWORDS = ["ticket", "билет", "eticket", "e-ticket", "itinerary", "booking confirmation",
                   "flight confirmation", "маршрут", "бронирование", "azerbaijan airlines",
                   "turkish airlines", "flydubai", "pegasus", "wizz", "lufthansa", "booking ref",
                   "flyarystan", "тарифный пакет", "номер билета", "electronic ticket", "маршрут-квитанция"]

SYSTEM_PROMPT = """Ты — ассистент по обработке авиабилетов. Пользователь присылает текст или изображение билета, либо команду для управления списком.

Верни ТОЛЬКО валидный JSON без markdown и блоков кода.

=== ДОБАВЛЕНИЕ БИЛЕТА ===
Если пользователь присылает данные билета:
{"action":"add","ticket":{"num":"номер","date":"YYYY-MM-DD","status":"Выписан","name":"Имя Пассажира","route":"Bak-Tbs","company":"","price":0,"currency":"AZN","cu":0,"ca":0},"missing":[]}

Правила добавления:
- status: "Выписан", "Изменён" или "Отменён". По умолчанию "Выписан"
- date: формат YYYY-MM-DD. Если не указана — сегодняшняя дата
- name: латиница, первая буква заглавная. Например: Ramazanov Elchin
- route: первая буква каждого сегмента заглавная. Например: Bak-Tbs-Bak
- price: берётся ТОЛЬКО из строки TOTAL. Например "TOTAL : AZN 164.70" -> price=164.70, currency=AZN. Если "TOTAL : EUR 210.00" -> price=210.00, currency=EUR. НЕ использовать AIR FARE, EQUIV FARE PAID или другие строки — только финальный TOTAL.
- currency: валюта из строки TOTAL — AZN, EUR, USD, RUB или KZT. По умолчанию AZN.
- cu: наша комиссия в AZN (если не указана — 0)
- ca: комиссия агента в AZN (если не указана — 0)
- company: компания-заказчик (если не указана — пустая строка)
- missing: список отсутствующих полей

=== ВОЗВРАТ БИЛЕТА ===
Если пользователь пишет "возврат", "это возврат", "верни X", "возврат X AZN":
{"action":"refund","name":"Имя Пассажира","route":"Bak-Tbs","amount":50}

=== УДАЛЕНИЕ БИЛЕТА ===
Если пользователь хочет удалить билет:
{"action":"delete","name":"Имя Пассажира","route":"Bak-Tbs"}

=== УДАЛЕНИЕ ВСЕГО ===
Если пользователь пишет "удали всё", "очисти всё":
{"action":"delete_all"}

=== ИЗМЕНЕНИЕ БИЛЕТА ===
{"action":"update","name":"Имя Пассажира","route":"Bak-Tbs","fields":{"поле":"новое значение"}}
Возможные поля: status, company, price, cu, ca, route, date
Если пишет "исправь цену X на Y" — fields:{"price":Y}
ВАЖНО: Если пользователь пишет только имя и комиссии (например "Salamov 25 наши 5 агента") — это ВСЕГДА update, НЕ add.
Пример: "Salamov Savkhan sco bak 25 наши, 5 агента" -> {"action":"update","name":"Salamov Savkhan","route":"Sco-Bak","fields":{"cu":25,"ca":5}}
Пример: "Ramazanov наша 15, агенту 5, компания Evrascon" -> {"action":"update","name":"Ramazanov","route":"","fields":{"cu":15,"ca":5,"company":"Evrascon"}}

=== ДРУГОЕ ===
Если это вопрос или непонятный текст: {"action":"chat","text":"твой ответ на русском"}"""

EMAIL_SYSTEM_PROMPT = """Ты — ассистент по обработке авиабилетов. Тебе передан текст письма с билетом.

Извлеки данные и верни ТОЛЬКО валидный JSON без markdown:
{"is_ticket":true,"ticket":{"num":"номер","date":"YYYY-MM-DD","status":"Выписан","name":"Имя Пассажира","route":"Bak-Tbs","company":"","price":0,"currency":"AZN"}}

Если это не билет: {"is_ticket":false}

Правила:
- is_ticket: true только если это реальный авиабилет с номером рейса и пассажиром
- num: номер билета или PNR
- date: дата вылета в формате YYYY-MM-DD
- name: имя пассажира латиницей, первая буква заглавная
- route: маршрут в формате Bak-Tbs или Bak-Tbs-Bak
- price: берётся ТОЛЬКО из строки TOTAL. Например "TOTAL : AZN 164.70" -> price=164.70, currency=AZN. Если "TOTAL : EUR 210.00" -> price=210.00, currency=EUR. НЕ использовать AIR FARE, EQUIV FARE PAID или другие строки — только финальный TOTAL. Если TOTAL не найден явно — ищи наибольшую итоговую сумму в письме.
- currency: валюта из строки TOTAL — AZN, EUR, USD, RUB или KZT. По умолчанию AZN.
- company: авиакомпания или компания-заказчик если указана"""


def get_cbar_rates():
    try:
        url = "https://www.cbar.az/currencies/today.xml"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as response:
            xml_data = response.read()
        root = ET.fromstring(xml_data)
        rates = {}
        for valute in root.findall(".//Valute"):
            code = valute.get("Code", "")
            nominal_el = valute.find("Nominal")
            value_el = valute.find("Value")
            if code and nominal_el is not None and value_el is not None:
                nominal = float(nominal_el.text)
                value = float(value_el.text.replace(",", "."))
                rates[code] = value / nominal
        return rates
    except Exception as e:
        logger.error(f"CBAR error: {e}")
        return {"USD": 1.70, "EUR": 1.87, "RUB": 0.019, "KZT": 0.0035}


def convert_to_azn(amount, currency, rates):
    if currency == "AZN" or not currency:
        return amount, 1.0
    rate = rates.get(currency.upper(), 1.0)
    return round(amount * rate, 2), rate


def load_tickets():
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT data FROM tickets ORDER BY (data->>'id')::bigint")
                return [dict(row['data']) for row in cur.fetchall()]
    except Exception as e:
        logger.error(f"load_tickets error: {e}")
        return []


def save_tickets(tickets):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM tickets")
                for t in tickets:
                    cur.execute(
                        "INSERT INTO tickets (id, data) VALUES (%s, %s)",
                        (t['id'], json.dumps(t, ensure_ascii=False))
                    )
            conn.commit()
    except Exception as e:
        logger.error(f"save_tickets error: {e}")


def load_seen_emails():
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT email_id FROM seen_emails")
                return set(row[0] for row in cur.fetchall())
    except Exception as e:
        logger.error(f"load_seen_emails error: {e}")
        return set()


def save_seen_emails(seen):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM seen_emails")
                for eid in seen:
                    cur.execute("INSERT INTO seen_emails (email_id) VALUES (%s)", (eid,))
            conn.commit()
    except Exception as e:
        logger.error(f"save_seen_emails error: {e}")


def fmt_date(d):
    if not d:
        return ""
    try:
        parts = d.split("-")
        return f"{parts[2]}.{parts[1]}.{parts[0]}"
    except:
        return d


def money(v):
    try:
        return f"{float(v):,.2f}"
    except:
        return "0.00"


def find_tickets(tickets, name=None, route=None):
    results = []
    for i, t in enumerate(tickets):
        match = True
        if name:
            n = name.upper().strip()
            tn = t.get("name", "").upper().strip()
            if n not in tn and tn not in n:
                match = False
        if route:
            r = route.upper().strip()
            tr = t.get("route", "").upper().strip()
            if r not in tr and tr not in r:
                match = False
        if match:
            results.append(i)
    return results


def decode_mime_words(s):
    if not s:
        return ""
    decoded = decode_header(s)
    result = []
    for part, enc in decoded:
        if isinstance(part, bytes):
            result.append(part.decode(enc or "utf-8", errors="ignore"))
        else:
            result.append(part)
    return " ".join(result)


def get_email_text(msg):
    import re

    def html_to_text(html):
        # Заменяем блочные теги на переносы строк, чтобы не склеивались слова
        html = re.sub(r'<(br|BR)[^>]*>', '\n', html)
        html = re.sub(r'</(td|TD|th|TH)>', ' ', html)
        html = re.sub(r'</(tr|TR|p|P|div|DIV|li|LI)>', '\n', html)
        html = re.sub(r'<[^>]+>', '', html)
        # Схлопываем множественные пробелы, но сохраняем переносы
        lines = [re.sub(r' +', ' ', line).strip() for line in html.split('\n')]
        return '\n'.join(l for l in lines if l)

    text = ""
    html_fallback = ""
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    text += payload.decode(part.get_content_charset() or "utf-8", errors="ignore")
            elif ctype == "text/html":
                payload = part.get_payload(decode=True)
                if payload:
                    html = payload.decode(part.get_content_charset() or "utf-8", errors="ignore")
                    html_fallback += html_to_text(html)
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            raw = payload.decode(msg.get_content_charset() or "utf-8", errors="ignore")
            if '<html' in raw.lower() or '<td' in raw.lower():
                text = html_to_text(raw)
            else:
                text = raw
    result = text if text.strip() else html_fallback
    return result[:6000]


def is_ticket_email(subject, text):
    combined = (subject + " " + text).lower()
    return any(kw.lower() in combined for kw in TICKET_KEYWORDS)


async def parse_email_with_claude(email_text, subject):
    prompt = f"Тема письма: {subject}\n\nТекст письма:\n{email_text}"
    response = await ANTHROPIC_CLIENT.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=1000,
        system=EMAIL_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = response.content[0].text.strip().replace("```json", "").replace("```", "").strip()
    return json.loads(raw)


def get_email_images(msg):
    """Извлекает изображения из вложений письма."""
    images = []
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype in ("image/jpeg", "image/png", "image/gif", "image/webp"):
                payload = part.get_payload(decode=True)
                if payload:
                    images.append((base64.standard_b64encode(payload).decode("utf-8"), ctype))
    return images


async def check_gmail(bot: Bot):
    if not GMAIL_USER or not GMAIL_PASSWORD:
        return
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(GMAIL_USER, GMAIL_PASSWORD)
        mail.select("inbox")
        _, data = mail.search(None, "UNSEEN")
        email_ids = data[0].split()
        if not email_ids:
            mail.logout()
            return
        seen = load_seen_emails()
        new_tickets = []
        for eid in email_ids[-20:]:
            eid_str = eid.decode()
            if eid_str in seen:
                continue
            _, msg_data = mail.fetch(eid, "(RFC822)")
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)
            subject = decode_mime_words(msg.get("Subject", ""))
            email_text = get_email_text(msg)
            images = get_email_images(msg)
            if not is_ticket_email(subject, email_text) and not images:
                seen.add(eid_str)
                continue
            try:
                # Если есть изображения — передаём первое в Claude вместе с текстом
                if images:
                    img_data, img_mime = images[0]
                    result = await parse_ticket_with_claude(
                        text=(f"Тема письма: {subject}\n\n{email_text}" if email_text.strip() else None),
                        image_data=img_data,
                        image_mime=img_mime
                    )
                    # parse_ticket_with_claude возвращает формат add, конвертируем в email формат
                    if result.get("action") == "add" and result.get("ticket"):
                        result = {"is_ticket": True, "ticket": result["ticket"]}
                    else:
                        result = {"is_ticket": False}
                else:
                    result = await parse_email_with_claude(email_text, subject)
                if result.get("is_ticket") and result.get("ticket"):
                    t = result["ticket"]
                    rates = get_cbar_rates()
                    price_orig = float(t.get("price", 0))
                    currency = t.get("currency", "AZN").upper()
                    price_azn, rate = convert_to_azn(price_orig, currency, rates)
                    ticket = {
                        "id": int(datetime.now().timestamp() * 1000),
                        "num": t.get("num", "—"),
                        "date": t.get("date", datetime.now().strftime("%Y-%m-%d")),
                        "status": t.get("status", "Выписан"),
                        "name": t.get("name", "—"),
                        "route": t.get("route", "—"),
                        "company": t.get("company", "—"),
                        "price_orig": price_orig,
                        "currency": currency,
                        "rate": rate,
                        "price_azn": price_azn,
                        "price": price_azn,
                        "cu": 0,
                        "ca": 0,
                        "owesUs": price_azn,
                        "owesAgent": price_azn,
                        "from_email": True
                    }
                    tickets = load_tickets()
                    tickets.append(ticket)
                    save_tickets(tickets)
                    new_tickets.append((ticket, subject))
            except Exception as e:
                logger.error(f"Email parse error: {e}")
            seen.add(eid_str)
        save_seen_emails(seen)
        mail.logout()
        for ticket, subject in new_tickets:
            currency_note = ""
            if ticket["currency"] != "AZN":
                currency_note = f"\n💱 {money(ticket['price_orig'])} {ticket['currency']} = {money(ticket['price_azn'])} AZN"
            msg_text = (
                f"📧 Новый билет из почты!\n"
                f"*{ticket['name']}* · {ticket['route']}\n"
                f"Дата: {fmt_date(ticket['date'])}\n"
                f"Номер: {ticket['num']}{currency_note}\n"
                f"Цена: {money(ticket['price_azn'])} AZN\n\n"
                f"⚠️ Укажите комиссии и компанию:\n"
                f"_наша комиссия 15, агенту 5, компания Evrascon_"
            )
            await bot.send_message(
                chat_id=ALLOWED_USER_ID,
                text=msg_text,
                parse_mode="Markdown"
            )
    except Exception as e:
        logger.error(f"Gmail check error: {e}")


async def parse_ticket_with_claude(text=None, image_data=None, image_mime=None):
    content = []
    if image_data:
        content.append({"type": "image", "source": {"type": "base64", "media_type": image_mime, "data": image_data}})
    if text:
        content.append({"type": "text", "text": text})
    elif image_data:
        content.append({"type": "text", "text": "Распознай данные билета с этого изображения"})
    if not content:
        content.append({"type": "text", "text": "Помогите с билетом"})
    response = await ANTHROPIC_CLIENT.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=1000,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": content}]
    )
    raw = response.content[0].text.strip().replace("```json", "").replace("```", "").strip()
    return json.loads(raw)


def generate_excel(tickets):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Билеты"
    headers = ["№", "Номер билета", "Дата", "Статус", "Пассажир", "Маршрут",
               "Компания", "Цена (ориг.)", "Валюта", "Курс", "Цена (AZN)",
               "Ком. наша", "Ком. агента", "Компания должна нам", "Мы должны агенту", "Возврат (AZN)"]
    header_fill = PatternFill(start_color="2C2C2A", end_color="2C2C2A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    col_widths = [4, 18, 12, 11, 24, 14, 20, 12, 8, 8, 12, 12, 13, 20, 20, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    status_colors = {"Выписан": "E1F5EE", "Изменён": "FAEEDA", "Отменён": "FCEBEB"}
    total_us = 0
    total_ag = 0
    for row_idx, t in enumerate(tickets, 2):
        price_azn = float(t.get("price_azn", t.get("price", 0)))
        cu = float(t.get("cu", 0))
        ca = float(t.get("ca", 0))
        owes_us = price_azn + cu + ca
        owes_ag = price_azn + ca
        total_us += owes_us
        total_ag += owes_ag
        refund = float(t.get("refund", 0))
        row_data = [
            row_idx - 1, t.get("num", ""), fmt_date(t.get("date", "")), t.get("status", ""),
            t.get("name", ""), t.get("route", ""), t.get("company", ""),
            float(t.get("price_orig", t.get("price", 0))), t.get("currency", "AZN"),
            float(t.get("rate", 1.0)), price_azn, cu, ca, owes_us, owes_ag,
            refund if refund else ""
        ]
        status = t.get("status", "")
        row_color = status_colors.get(status, "FFFFFF")
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col in [8, 10, 11, 12, 13, 14, 15] or (col == 16 and val != ""):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            if col == 4:
                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
    total_row = len(tickets) + 2
    ws.cell(row=total_row, column=7, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=total_row, column=14, value=total_us).number_format = '#,##0.00'
    ws.cell(row=total_row, column=14).font = Font(bold=True, color="0F6E56")
    ws.cell(row=total_row, column=15, value=total_ag).number_format = '#,##0.00'
    ws.cell(row=total_row, column=15).font = Font(bold=True, color="854F0B")
    total_refund = sum(float(t.get("refund", 0)) for t in tickets)
    if total_refund:
        ws.cell(row=total_row, column=16, value=total_refund).number_format = '#,##0.00'
        ws.cell(row=total_row, column=16).font = Font(bold=True, color="A32D2D")
    filename = f"/tmp/tickets_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    return filename


def is_allowed(update: Update):
    if ALLOWED_USER_ID == 0:
        return True
    return update.effective_user.id == ALLOWED_USER_ID


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    keyboard = [[KeyboardButton("📊 Отчёт Excel")], [KeyboardButton("📋 Список билетов"), KeyboardButton("🗑 Очистить всё")]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    email_status = f"📧 Почта: {GMAIL_USER}" if GMAIL_USER else "📧 Почта: не подключена"
    await update.message.reply_text(
        f"Привет! Я помогаю собирать авиабилеты в отчёт.\n\n"
        f"{email_status} (слежу за почтой мгновенно)\n\n"
        "Отправьте мне:\n"
        "• Фото или скриншот билета\n"
        "• Или текст с данными билета\n\n"
        "Управление:\n"
        "• _удали Ramazanov Elchin Bak-Tbs_\n"
        "• _измени Ramazanov статус Отменён_\n"
        "• _исправь цену Ramazanov 180_\n"
        "• _возврат 50 azn Ramazanov_\n"
        "• _удали всё_",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    await update.message.reply_text("Обрабатываю билет, подождите...")
    photo = update.message.photo[-1]
    file = await photo.get_file()
    file_bytes = await file.download_as_bytearray()
    image_data = base64.standard_b64encode(file_bytes).decode("utf-8")
    caption = update.message.caption or ""
    try:
        result = await parse_ticket_with_claude(text=caption if caption else None, image_data=image_data, image_mime="image/jpeg")
        await process_result(update, context, result)
    except Exception as e:
        logger.error(f"Error: {e}")
        await update.message.reply_text("Не смог распознать билет. Попробуйте отправить текстом.")


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    text = update.message.text
    if text == "📊 Отчёт Excel":
        await send_report(update, context)
        return
    elif text == "📋 Список билетов":
        await list_tickets(update, context)
        return
    elif text == "🗑 Очистить всё":
        save_tickets([])
        await update.message.reply_text("Список очищен.")
        return
    await update.message.reply_text("Обрабатываю...")
    try:
        result = await parse_ticket_with_claude(text=text)
        await process_result(update, context, result)
    except Exception as e:
        logger.error(f"Error: {e}")
        await update.message.reply_text("Не смог разобрать. Попробуйте другой формат.")


async def process_result(update: Update, context: ContextTypes.DEFAULT_TYPE, result: dict):
    action = result.get("action")

    if action == "refund":
        tickets = load_tickets()
        name = result.get("name", "")
        route = result.get("route", "")
        amount = float(result.get("amount", 0))
        indices = find_tickets(tickets, name=name, route=route)
        if not indices:
            await update.message.reply_text("❌ Билет не найден. Проверьте имя или маршрут.")
            return
        for i in indices:
            tickets[i]["status"] = "Отменён"
            tickets[i]["refund"] = -abs(amount)
        save_tickets(tickets)
        lines = [f"↩️ Возврат оформлен: *-{money(amount)} AZN*"]
        for i in indices:
            t = tickets[i]
            lines.append(f"• {t['name']} · {t['route']} · {fmt_date(t['date'])}\n  Статус → Отменён")
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
        return

    if action == "delete_all":
        save_tickets([])
        await update.message.reply_text("🗑 Все билеты удалены.")
        return

    if action == "delete":
        tickets = load_tickets()
        name = result.get("name", "")
        route = result.get("route", "")
        indices = find_tickets(tickets, name=name, route=route)
        if not indices:
            await update.message.reply_text("❌ Билет не найден. Проверьте имя или маршрут.")
            return
        deleted = [tickets[i] for i in indices]
        tickets = [t for i, t in enumerate(tickets) if i not in indices]
        save_tickets(tickets)
        lines = [f"🗑 Удалено {len(deleted)} билет(ов):"]
        for t in deleted:
            lines.append(f"• {t['name']} · {t['route']} · {fmt_date(t['date'])}")
        await update.message.reply_text("\n".join(lines))
        return

    if action == "update":
        tickets = load_tickets()
        name = result.get("name", "")
        route = result.get("route", "")
        fields = result.get("fields", {})
        indices = find_tickets(tickets, name=name, route=route)
        if not indices:
            await update.message.reply_text("❌ Билет не найден. Проверьте имя или маршрут.")
            return
        if not fields:
            await update.message.reply_text("❌ Не указано что именно изменить.")
            return
        rates = get_cbar_rates()
        for i in indices:
            for field, value in fields.items():
                tickets[i][field] = value
            if "price" in fields:
                currency = tickets[i].get("currency", "AZN")
                new_price_azn, rate = convert_to_azn(float(fields["price"]), currency, rates)
                tickets[i]["price_orig"] = float(fields["price"])
                tickets[i]["price_azn"] = new_price_azn
                tickets[i]["price"] = new_price_azn
                tickets[i]["rate"] = rate
            price_azn = float(tickets[i].get("price_azn", tickets[i].get("price", 0)))
            cu = float(tickets[i].get("cu", 0))
            ca = float(tickets[i].get("ca", 0))
            tickets[i]["owesUs"] = price_azn + cu + ca
            tickets[i]["owesAgent"] = price_azn + ca
        save_tickets(tickets)
        field_names = {"status": "статус", "company": "компания", "price": "цена", "cu": "ком. наша", "ca": "ком. агента", "route": "маршрут", "date": "дата"}
        changed = ", ".join([f"{field_names.get(k, k)}: {v}" for k, v in fields.items()])
        lines = [f"✏️ Изменено {len(indices)} билет(ов): {changed}"]
        for i in indices:
            t = tickets[i]
            lines.append(f"• {t['name']} · {t['route']} · {fmt_date(t['date'])}")
        await update.message.reply_text("\n".join(lines))
        return

    if action == "add" and result.get("ticket"):
        t = result["ticket"]

        # Если цена 0, но есть комиссии/компания — скорее всего это обновление существующего билета
        price_orig_check = float(t.get("price", 0))
        cu_check = float(t.get("cu", 0))
        ca_check = float(t.get("ca", 0))
        company_check = t.get("company", "")
        if price_orig_check == 0 and (cu_check or ca_check or company_check):
            existing = load_tickets()
            name_check = t.get("name", "")
            route_check = t.get("route", "")
            indices = find_tickets(existing, name=name_check, route=route_check if route_check and route_check != "—" else None)
            if indices:
                fields = {}
                if cu_check: fields["cu"] = cu_check
                if ca_check: fields["ca"] = ca_check
                if company_check and company_check != "—": fields["company"] = company_check
                for i in indices:
                    existing[i].update(fields)
                    price_azn = float(existing[i].get("price_azn", existing[i].get("price", 0)))
                    existing[i]["owesUs"] = price_azn + float(existing[i].get("cu", 0)) + float(existing[i].get("ca", 0))
                    existing[i]["owesAgent"] = price_azn + float(existing[i].get("ca", 0))
                save_tickets(existing)
                field_names = {"cu": "ком. наша", "ca": "ком. агента", "company": "компания"}
                changed = ", ".join([f"{field_names.get(k,k)}: {v}" for k,v in fields.items()])
                lines = [f"✏️ Обновил {len(indices)} билет(ов): {changed}"]
                for i in indices:
                    tk = existing[i]
                    line = (f"• {tk['name']} · {tk['route']} · {fmt_date(tk['date'])}\n"
                            f"  💰 Цена: {money(tk.get('price_azn', tk.get('price',0)))} AZN · Ком. наша: {money(tk.get('cu',0))} · Ком. агента: {money(tk.get('ca',0))}\n"
                            f"  🟢 Нам: {money(tk['owesUs'])} AZN · 🟡 Агенту: {money(tk['owesAgent'])} AZN")
                    lines.append(line)
                await update.message.reply_text("\n".join(lines))
                return

        price_orig = float(t.get("price", 0))
        currency = t.get("currency", "AZN").upper()
        cu = float(t.get("cu", 0))
        ca = float(t.get("ca", 0))
        rates = get_cbar_rates()
        price_azn, rate = convert_to_azn(price_orig, currency, rates)
        owes_us = price_azn + cu + ca
        owes_ag = price_azn + ca
        ticket = {
            "id": int(datetime.now().timestamp() * 1000),
            "num": t.get("num", "—"),
            "date": t.get("date", datetime.now().strftime("%Y-%m-%d")),
            "status": t.get("status", "Выписан"),
            "name": t.get("name", "—"),
            "route": t.get("route", "—"),
            "company": t.get("company", "—"),
            "price_orig": price_orig,
            "currency": currency,
            "rate": rate,
            "price_azn": price_azn,
            "price": price_azn,
            "cu": cu,
            "ca": ca,
            "owesUs": owes_us,
            "owesAgent": owes_ag
        }
        tickets = load_tickets()
        tickets.append(ticket)
        save_tickets(tickets)
        missing = result.get("missing", [])
        missing_note = ""
        if missing:
            labels = {"cu": "наша комиссия", "ca": "комиссия агента", "company": "компания", "price": "цена"}
            missing_names = [labels.get(m, m) for m in missing]
            missing_note = f"\n\n⚠️ Не нашёл: {', '.join(missing_names)} — поставил 0"
        currency_note = ""
        if currency != "AZN":
            currency_note = f"\n💱 {money(price_orig)} {currency} × {rate:.4f} = {money(price_azn)} AZN"
        msg = (
            f"✅ Добавил билет *{ticket['num']}*\n"
            f"👤 {ticket['name']}\n"
            f"✈️ {ticket['route']} · {fmt_date(ticket['date'])}\n"
            f"🏢 Компания: {ticket['company']}"
            f"{currency_note}\n"
            f"💰 Цена: {money(price_azn)} AZN · Ком. наша: {money(cu)} · Ком. агента: {money(ca)}\n\n"
            f"🟢 Компания должна нам: *{money(owes_us)} AZN*\n"
            f"🟡 Мы должны агенту: *{money(owes_ag)} AZN*"
            f"{missing_note}"
        )
        await update.message.reply_text(msg, parse_mode="Markdown")
        return

    if action == "chat":
        await update.message.reply_text(result.get("text", "Не понял запрос."))
        return

    await update.message.reply_text("Не смог распознать. Попробуйте другой формат.")


async def send_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tickets = load_tickets()
    if not tickets:
        await update.message.reply_text("Нет билетов для отчёта.")
        return
    await update.message.reply_text("Формирую Excel...")
    filename = generate_excel(tickets)
    total_us = sum(t.get("owesUs", 0) for t in tickets)
    total_ag = sum(t.get("owesAgent", 0) for t in tickets)
    with open(filename, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=filename,
            caption=f"📊 Отчёт · {len(tickets)} билетов\n🟢 Должны нам: {money(total_us)} AZN\n🟡 Мы агентам: {money(total_ag)} AZN"
        )
    os.remove(filename)


async def list_tickets(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tickets = load_tickets()
    if not tickets:
        await update.message.reply_text("Список пуст.")
        return
    lines = [f"📋 *Билеты ({len(tickets)} шт.)*\n"]
    for i, t in enumerate(tickets[-10:], 1):
        price_azn = t.get("price_azn", t.get("price", 0))
        curr = t.get("currency", "AZN")
        curr_note = f" ({money(t.get('price_orig', price_azn))} {curr})" if curr != "AZN" else ""
        email_note = " 📧" if t.get("from_email") else ""
        lines.append(f"{i}. *{t['num']}* — {t['name']}{email_note}\n   {t['route']} · {fmt_date(t['date'])} · {t.get('company','')}\n   {money(price_azn)} AZN{curr_note}")
    if len(tickets) > 10:
        lines.append(f"\n_...и ещё {len(tickets)-10} билетов_")
    total_us = sum(t.get("owesUs", 0) for t in tickets)
    total_ag = sum(t.get("owesAgent", 0) for t in tickets)
    lines.append(f"\n🟢 Итого нам: *{money(total_us)} AZN*\n🟡 Итого агентам: *{money(total_ag)} AZN*")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def email_check_job(bot: Bot):
    """Следит за почтой через IMAP IDLE — реагирует мгновенно при новом письме."""
    while True:
        if not GMAIL_USER or not GMAIL_PASSWORD:
            await asyncio.sleep(60)
            continue
        try:
            loop = asyncio.get_event_loop()

            def imap_idle():
                mail = imaplib.IMAP4_SSL("imap.gmail.com")
                mail.login(GMAIL_USER, GMAIL_PASSWORD)
                mail.select("inbox")
                tag = mail._new_tag().decode()
                mail.send(f"{tag} IDLE\r\n".encode())
                mail.sock.settimeout(1680)
                try:
                    while True:
                        line = mail.readline().decode(errors="ignore").strip()
                        if not line:
                            continue
                        logger.info(f"IMAP IDLE: {line}")
                        if "EXISTS" in line or "RECENT" in line:
                            break
                except Exception:
                    pass
                finally:
                    try:
                        mail.send(b"DONE\r\n")
                        mail.logout()
                    except Exception:
                        pass

            logger.info("IMAP IDLE: ожидаем новые письма...")
            await loop.run_in_executor(None, imap_idle)
            logger.info("IMAP IDLE: получено уведомление, проверяем почту...")
            await check_gmail(bot)
        except Exception as e:
            logger.error(f"IMAP IDLE error: {e}")
            await asyncio.sleep(30)


async def post_init(application):
    application.create_task(email_check_job(application.bot))


def main():
    init_db()
    logger.info("Database initialized")
    app = Application.builder().token(TELEGRAM_TOKEN).post_init(post_init).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    logger.info("Bot started with Gmail monitoring...")
    app.run_polling()


if __name__ == "__main__":
    main()
